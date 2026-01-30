import openpyxl

# Make new excel file
wb = openpyxl.Workbook()

# Select activated current sheet
ws = wb.active

# Change sheet name
ws.title = 'Sheet_by_automation_code'

# Make new sheet
ws = wb.create_sheet("26-01-31")

# Delete a sheet
del wb['Sheet_by_automation_code']

# Print out all sheet(s)
print(wb.sheetnames)

# Save the file
wb.save("Automated_Excel_File.xlsx")