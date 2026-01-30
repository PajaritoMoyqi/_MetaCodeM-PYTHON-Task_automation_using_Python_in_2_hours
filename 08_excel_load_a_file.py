import openpyxl

# Load a excel file
wb = openpyxl.load_workbook('Automated_Excel_File.xlsx')

# Select activated current sheet
ws = wb.active

# Add data(s)
ws['A1'] = 'Date'
ws['B1'] = 'Name'
ws['C1'] = 'Price'
ws['D1'] = 'Amount'
ws['E1'] = 'Total'

ws.cell(row=2, column=1, value="2026-01-31")
ws.cell(row=2, column=2, value="Pencil")
ws.cell(row=2, column=3, value=300)
ws.cell(row=2, column=4, value=120)
ws.cell(row=2, column=5, value="=C2*D2")

ws.append(["2026-01-31", "TV", 300000, 1, "=C3*D3"])

# Change data(s)
ws["C2"] = 400
ws["D2"] = 1000

# Remove data(s)
del ws["A2"]

# Save the file
wb.save("Automated_Excel_File.xlsx")